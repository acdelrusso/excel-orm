from __future__ import annotations

import re
from dataclasses import dataclass
from typing import Any, TypeVar

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet

from src.column import Column

M = TypeVar("M")


def _camel_to_snake(name: str) -> str:
    s1 = re.sub("(.)([A-Z][a-z]+)", r"\1_\2", name)
    s2 = re.sub("([a-z0-9])([A-Z])", r"\1_\2", s1)
    return s2.lower()


def _pluralize(s: str) -> str:
    # keep deliberately simple; can be swapped for inflect later
    if s.endswith("s"):
        return s
    return s + "s"


def _repo_name_for_model(model: type[Any]) -> str:
    return _pluralize(_camel_to_snake(model.__name__))


def _display_name_for_model(model: type[Any]) -> str:
    # "manufacturing_plants" -> "Manufacturing Plants"
    return _repo_name_for_model(model).replace("_", " ").title()


def _get_model_columns(model: type[Any]) -> list[Column[Any]]:
    """
    Return declared Column descriptors on the class in definition order.
    Uses __annotations__ order as the source of truth.
    """
    cols: list[Column[Any]] = []
    ann = getattr(model, "__annotations__", {})
    for field_name in ann:
        v = getattr(model, field_name, None)
        if isinstance(v, Column):
            cols.append(v)
    return cols


def _normalize_header(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _row_is_blank(values: list[Any]) -> bool:
    return all(_normalize_header(v) == "" for v in values)


def _instantiate_model[M](model: type[M]) -> M:
    obj = model.__new__(model)
    obj._values = {}
    # defaults
    for col in _get_model_columns(model):
        if col.name is None:
            raise RuntimeError("Column __set_name__ did not run.")
        obj._values[col.name] = col.spec.default
    return obj


class Repository(list[M]):
    def all(self) -> list[M]:
        return list(self)


@dataclass(frozen=True)
class SheetSpec:
    name: str
    models: list[type[Any]]

    title_row: int = 1
    header_row: int = 2
    data_start_row: int = 3

    template_table_gap: int = 2


class ExcelFile:
    def __init__(self, *, sheets: list[SheetSpec]):
        self.sheets = sheets

        self._repos: dict[type[Any], Repository[Any]] = {}

        for sheet in sheets:
            for model in sheet.models:
                repo_name = _repo_name_for_model(model)
                if hasattr(self, repo_name):
                    raise ValueError(
                        f"Duplicate repo name '{repo_name}' for model {model.__name__}"
                    )
                repo = Repository()
                self._repos[model] = repo
                setattr(self, repo_name, repo)

    def generate_template(self, filename: str) -> None:
        wb = Workbook()
        # remove default sheet if we will create our own named sheets
        default_ws = wb.active
        if len(self.sheets) > 0:
            wb.remove(default_ws)

        for sheet in self.sheets:
            ws = wb.create_sheet(title=sheet.name)
            self._write_sheet_template(ws, sheet)

        wb.save(filename)

    def _write_sheet_template(self, ws: Worksheet, spec: SheetSpec) -> None:
        current_col = 1  # 1-based index

        title_font = Font(bold=True)
        title_alignment = Alignment(horizontal="center", vertical="center")

        header_font = Font(bold=True)

        for model in spec.models:
            cols = _get_model_columns(model)
            headers = [c.spec.header or c.name for c in cols]
            width = len(headers)

            start_col = current_col
            end_col = current_col + width - 1

            # ---- merged title row ----
            ws.merge_cells(
                start_row=spec.title_row,
                start_column=start_col,
                end_row=spec.title_row,
                end_column=end_col,
            )
            title_cell = ws.cell(
                row=spec.title_row, column=start_col, value=_display_name_for_model(model)
            )
            title_cell.font = title_font
            title_cell.alignment = title_alignment

            for j, h in enumerate(headers):
                c = start_col + j
                cell = ws.cell(row=spec.header_row, column=c, value=h)
                cell.font = header_font

                col_letter = ws.cell(row=spec.header_row, column=c).column_letter
                ws.column_dimensions[col_letter].width = max(12, min(40, len(str(h)) + 4))

            current_col = end_col + 1 + spec.template_table_gap

    def load_data(self, filename: str) -> None:
        wb = load_workbook(filename=filename, data_only=True)

        # clear old data
        for repo in self._repos.values():
            repo.clear()

        for sheet_spec in self.sheets:
            if sheet_spec.name not in wb.sheetnames:
                raise ValueError(f"Workbook missing sheet '{sheet_spec.name}'")
            ws = wb[sheet_spec.name]
            self._parse_sheet(ws, sheet_spec)

    def _parse_sheet(self, ws: Worksheet, spec: SheetSpec) -> None:
        for model in spec.models:
            found = self._find_header(ws, spec, model)
            if found is None:
                continue

            _, start_col = found
            cols = _get_model_columns(model)
            width = len(cols)

            repo: Repository[Any] = self._repos[model]

            r = spec.data_start_row
            while r <= ws.max_row:
                row_vals = [ws.cell(row=r, column=start_col + j).value for j in range(width)]
                if _row_is_blank(row_vals):
                    break

                # excludes (raw-value based)
                if any(
                    col.spec.excludes and row_vals[i] in col.spec.excludes
                    for i, col in enumerate(cols)
                ):
                    r += 1
                    continue

                obj = _instantiate_model(model)
                for i, col in enumerate(cols):
                    raw = row_vals[i]
                    parsed = col.parse_cell(raw)
                    setattr(obj, col.name, parsed)

                validate = getattr(obj, "validate", None)
                if callable(validate):
                    validate()

                repo.append(obj)
                r += 1

    def _find_header(
        self, ws: Worksheet, spec: SheetSpec, model: type[Any]
    ) -> tuple[int, int] | None:
        cols = _get_model_columns(model)
        expected = [_normalize_header(c.spec.header) for c in cols]
        if not expected:
            return None

        r = spec.header_row
        width = len(expected)
        max_c = ws.max_column or 0

        for start_col in range(1, max_c - width + 2):
            actual = [
                _normalize_header(ws.cell(row=r, column=start_col + j).value) for j in range(width)
            ]
            if actual == expected:
                return (r, start_col)

        return None
