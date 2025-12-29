from __future__ import annotations

import pytest
from openpyxl import load_workbook


def test_generate_template_creates_sheet_and_titles_and_headers(tmp_path, excel_file):
    out = tmp_path / "template.xlsx"
    excel_file.generate_template(str(out))

    wb = load_workbook(out)
    assert "Cars" in wb.sheetnames
    ws = wb["Cars"]

    # Cars block: title row merged across A..C, title value "Cars"
    assert ws["A1"].value == "Cars"
    merged = [str(rng) for rng in ws.merged_cells.ranges]
    assert "A1:C1" in merged

    # Cars headers on row 2
    assert ws["A2"].value == "Make"
    assert ws["B2"].value == "Model"
    assert ws["C2"].value == "Year"

    # ManufacturingPlant block starts after 2-col gap: Cars ends at C, so start at F
    assert ws["F1"].value == "Manufacturing Plants"
    assert "F1:G1" in merged
    assert ws["F2"].value == "Factory Name"
    assert ws["G2"].value == "Location"


def test_load_data_end_to_end(tmp_path, excel_file):
    """
    Create a workbook that matches the template layout, fill a few rows,
    load it, and verify repositories.
    """
    out = tmp_path / "data.xlsx"
    excel_file.generate_template(str(out))

    # Fill data
    from openpyxl import load_workbook

    wb = load_workbook(out)
    ws = wb["Cars"]

    # Cars data at row 3+
    ws["A3"].value = "Toyota"
    ws["B3"].value = "Camry"
    ws["C3"].value = 2020

    ws["A4"].value = "Honda"
    ws["B4"].value = "Civic"
    ws["C4"].value = "2019"  # string should parse to int

    # Plants data at row 3+ (F,G)
    ws["F3"].value = "Plant 1"
    ws["G3"].value = "NJ"

    ws["F4"].value = "Plant 2"
    ws["G4"].value = "PA"

    wb.save(out)

    # Load
    excel_file.load_data(str(out))

    cars = excel_file.cars.all()
    plants = excel_file.manufacturing_plants.all()

    assert len(cars) == 2
    assert cars[0].make == "Toyota"
    assert cars[0].model == "Camry"
    assert cars[0].year == 2020
    assert cars[1].year == 2019

    assert len(plants) == 2
    assert plants[0].name == "Plant 1"
    assert plants[0].location == "NJ"


def test_load_data_missing_sheet_raises(tmp_path, excel_file):
    # Create a workbook with a different sheet name
    from openpyxl import Workbook

    p = tmp_path / "wrong.xlsx"
    wb = Workbook()
    wb.active.title = "NotCars"
    wb.save(p)

    with pytest.raises(ValueError):
        excel_file.load_data(str(p))
